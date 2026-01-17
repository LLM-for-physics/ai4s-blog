#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
统计所有学生的作业分数并导出到Excel表格
逻辑与 ScoreQuery.vue 保持一致
"""

import csv
from pathlib import Path
from typing import Dict, List, Optional
import yaml
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("需要安装 openpyxl 库，正在安装...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill


# 配置 - 与 Vue 组件保持一致
BASE_PATH = Path("docs/public/score")
SERVERS = ['58', '132', '197']
MAX_ASSIGNMENTS = 7
LECTURE_SCORE = 30  # 讲座课固定分数

# 各作业满分 - 与 Vue 的 getFullScore 函数一致
def get_full_score(assignment_num: int) -> int:
    if assignment_num == 5:
        return 5
    if assignment_num == 6:
        return 5
    if assignment_num == 7:
        return 20
    return 10


def parse_csv(csv_path: Path) -> List[Dict]:
    """解析CSV文件"""
    if not csv_path.exists():
        return []
    
    try:
        # 使用 utf-8-sig 编码自动处理 BOM
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            return list(reader)
    except Exception as e:
        print(f"解析CSV文件失败 {csv_path}: {e}")
        return []


def read_text_file(file_path: Path) -> Optional[str]:
    """读取文本文件内容"""
    if not file_path.exists():
        return None
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read().strip()
    except Exception:
        return None


def parse_score_text(score_text: str) -> Optional[float]:
    """
    解析分数文本，支持多种格式：
    - 纯数字: "10" -> 10.0
    - 带满分: "18.5/20" -> 18.5
    """
    score_text = score_text.strip()
    if '/' in score_text:
        # 格式: 分数/满分
        parts = score_text.split('/')
        try:
            return float(parts[0].strip())
        except ValueError:
            return None
    else:
        try:
            return float(score_text)
        except ValueError:
            return None


def fetch_score(server: str, student_id: str, assignment_num: int) -> Optional[float]:
    """
    读取分数文件 - 与 Vue 的逻辑完全一致
    先尝试 {num}-score.txt，再尝试 {num}_score.txt
    """
    student_dir = BASE_PATH / server / f"stu{student_id}"
    
    # 尝试两种文件命名格式：横线和下划线
    score_url1 = student_dir / f"{assignment_num}-score.txt"
    score_url2 = student_dir / f"{assignment_num}_score.txt"
    
    score_text = read_text_file(score_url1)
    if score_text is None:
        score_text = read_text_file(score_url2)
    
    if score_text is not None:
        return parse_score_text(score_text)
    return None


def load_extra_yaml() -> Dict:
    """加载extra.yaml中的图书馆agent比赛信息"""
    yaml_path = BASE_PATH / "extra.yaml"
    if not yaml_path.exists():
        return {}
    
    try:
        with open(yaml_path, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f)
        return data
    except Exception as e:
        print(f"加载extra.yaml失败: {e}")
        return {}


def check_agent_project(extra_data: Dict, student_id: str) -> Optional[Dict]:
    """检查图书馆agent比赛项目 - 与 Vue 的 checkAgentProject 函数一致"""
    if not extra_data or '图书馆 agent 比赛' not in extra_data:
        return None
    
    agent_data = extra_data['图书馆 agent 比赛']
    student_key = f"stu{student_id}"
    
    if student_key in agent_data:
        return {
            'score': agent_data[student_key].get('score'),
            'substitute_homework': agent_data[student_key].get('substitute_homework')
        }
    
    return None


def collect_all_students() -> Dict[str, Dict]:
    """
    收集所有学生信息 - 模拟 Vue 的查询逻辑
    返回格式: {学号: {'姓名': str, '分数': [7个分数或None]}}
    """
    students = {}
    extra_data = load_extra_yaml()
    
    # 遍历所有服务器和作业，收集学生信息
    for server in SERVERS:
        for assignment_num in range(1, MAX_ASSIGNMENTS + 1):
            csv_path = BASE_PATH / server / f"assignment{assignment_num}_update_log.csv"
            csv_data = parse_csv(csv_path)
            
            for row in csv_data:
                student_id = row.get('学号', '').strip()
                student_name = row.get('姓名', '').strip()
                
                if not student_id:
                    continue
                
                # 初始化学生信息
                if student_id not in students:
                    students[student_id] = {
                        '姓名': student_name,
                        '分数': [None] * MAX_ASSIGNMENTS
                    }
                
                # 如果该作业还没有分数，尝试读取
                if students[student_id]['分数'][assignment_num - 1] is None:
                    score = fetch_score(server, student_id, assignment_num)
                    if score is not None:
                        students[student_id]['分数'][assignment_num - 1] = score
    
    # 应用图书馆agent比赛的替代分数
    for student_id in students:
        agent_project = check_agent_project(extra_data, student_id)
        if agent_project and agent_project['substitute_homework']:
            hw_num = agent_project['substitute_homework']
            if 1 <= hw_num <= MAX_ASSIGNMENTS:
                students[student_id]['分数'][hw_num - 1] = agent_project['score']
    
    return students


def export_to_excel(students: Dict[str, Dict], output_file: str = "学生作业成绩统计.xlsx"):
    """导出到Excel"""
    print(f"导出到Excel文件: {output_file}")
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生成绩统计"
    
    # 设置表头
    headers = ['学号', '姓名']
    for i in range(1, MAX_ASSIGNMENTS + 1):
        headers.append(f'作业{i}')
    headers.extend(['讲座课', '总分'])
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # 按学号排序
    sorted_students = sorted(students.items(), key=lambda x: x[0])
    
    # 写入数据
    for row_idx, (student_id, info) in enumerate(sorted_students, 2):
        ws.cell(row=row_idx, column=1, value=student_id)
        ws.cell(row=row_idx, column=2, value=info['姓名'])
        
        # 写入各作业分数
        total = 0
        for col_idx, score in enumerate(info['分数'], 3):
            # None 表示未评分，显示为空；0 是实际的0分
            display_score = score if score is not None else ''
            cell = ws.cell(row=row_idx, column=col_idx, value=display_score)
            cell.alignment = Alignment(horizontal='center')
            if score is not None:
                total += score
        
        # 写入讲座课分数
        cell = ws.cell(row=row_idx, column=3 + MAX_ASSIGNMENTS, value=LECTURE_SCORE)
        cell.alignment = Alignment(horizontal='center')
        total += LECTURE_SCORE
        
        # 写入总分
        cell = ws.cell(row=row_idx, column=3 + MAX_ASSIGNMENTS + 1, value=total)
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    for col in range(3, 3 + MAX_ASSIGNMENTS + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 10
    
    # 保存文件
    wb.save(output_file)
    print(f"成功导出 {len(students)} 个学生的成绩到 {output_file}")


def main():
    print("开始收集学生成绩...")
    students = collect_all_students()
    print(f"共收集到 {len(students)} 个学生")
    
    export_to_excel(students)
    print("完成！")


if __name__ == "__main__":
    main()
